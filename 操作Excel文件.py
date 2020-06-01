import openpyxl
# 新建文件
wbk = openpyxl.Workbook()
wbk.remove(wbk.worksheets[0])
sht = wbk.create_sheet('General')
# 打开Excel文件
wbk = openpyxl.load_workbook(pic_excel_path + '\\' + file_name)
wbk = openpyxl.load_workbook(report_folder + '\\' + file, data_only=True)  # data_only=True => 读取公式的结果值而非公式本身
sht = wbk.worksheets[0] # 根据表格的次序
sht = wbk['abc']    # 根据表格的名字
# 遍历表单名字
for sht_name in wbk.sheetnames:
    sht = wbk[sht_name]
    if sht_name == 'Mode6变量':
        pass


# 保存Excel文件
wbk.save(export_result_excel_path)

# 修改Excel内容
sht.cell(row=1, column=y).value = 111
# 内容为空
sht.cell(row=1, column=y).value = None
# 调整列宽
sht.column_dimensions[changeNumToChar(toBigChar=y)].width = 20
#  条件格式
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
sht_db.conditional_formatting.add(cell_char, FormulaRule(formula=['ISBLANK(%s)'%(cell_char)], stopIfTrue=True, fill=yellow_fill))
#  填充颜色
pink_fill = PatternFill(start_color='FABF8F', end_color='FABF8F', fill_type='solid')
sht_db.cell(row=x_db, column=1 + i).fill = pink_fill
# 获取行数
x = len(tuple(sht.rows))

# 遍历数据
for i, value in enumerate(sht.values):
    print(value)
    if value[1] in [0, 1]:
        tag_list.append(value[1])
        amplitude_list.append(value[2])
        square_list.append(value[3])